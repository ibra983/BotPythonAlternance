import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import openpyxl
import schedule
import time
import os

def envoi_mail():
    # Paramètres SMTP
    smtp_server = 'smtp.office365.com'
    port = 587
    sender_email = 'monmail'
    app_password = 'mdpappli'

    # Charger le fichier Excel
    wb = openpyxl.load_workbook(r'C:\chemin\vers\mon\fichier\ListeEntreprises.xlsx')
    sheet = wb.active


    # Ensemble pour stocker les adresses e-mail uniques
    email_addresses = set()

    for row in sheet.iter_rows(values_only=True):
        for cell_value in row:
            if isinstance(cell_value, str) and '@' in cell_value:
                email_addresses.add(cell_value.strip())

    # chercher d'abord les adresses e-mail déjà contactées 
    # depuis le fichier en question 
    # pr savoir si on doit les contacter ou non
    already_contacted_file = 'maildejacontacte.txt'
    if os.path.exists(already_contacted_file):
        with open(already_contacted_file, 'r') as file:
            already_contacted_emails = set(file.read().splitlines())
    else:
        already_contacted_emails = set()

    # Établir une connexion avec le serveur SMTP
    server = smtplib.SMTP(smtp_server, port)
    server.starttls()

    # Se co au compte Outlook
    server.login(sender_email, app_password)

    # Parcourt les lignes du fichier et ajouter les adresses e-mail à l'ensemble
    for row in sheet.iter_rows(values_only=True):
        for cell_value in row:
            if isinstance(cell_value, str) and '@' in cell_value:
                email_addresses.add(cell_value.strip())




    # Parcourt les adresses e-mail uniques et envoyer les e-mails
    for recipient_email in email_addresses:
        if recipient_email not in already_contacted_emails:
            # Créer le message
            message = MIMEMultipart()
            message['From'] = sender_email
            message['To'] = recipient_email
            message['Subject'] = 'Candidature spontanée : rejoindre votre équipe informatique'
            body = 'corps du message'
            message.attach(MIMEText(body, 'plain', 'utf-8'))

                # y mettre les pièces jointes ( ici mon CV video )
            filename1 = r'C:\chemin\vers\mon\fichier/.mp4'

            with open(filename1, 'rb') as file1:
                pdf_content1 = file1.read()

            part1 = MIMEApplication(pdf_content1, Name=os.path.basename(filename1))
            part1.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(filename1)}"')
            message.attach(part1)

            # Envoyer l'e-mail
            server.sendmail(sender_email, recipient_email, message.as_string().encode('utf-8'))

            # Ajouter l'adresse e-mail aux adresses déjà contactées
            already_contacted_emails.add(recipient_email)

            with open(already_contacted_file, 'a') as file:
                file.write(recipient_email + '\n')
            
            print(f'E-mail envoyé / SUCCES : {recipient_email}' )
        else:
            print(f'E-mail déjà envoyé / ECHEC : {recipient_email}')
envoi_mail()     
schedule.every(0.05).seconds.do(envoi_mail)
while True:
    schedule.run_pending()
    time.sleep(1)